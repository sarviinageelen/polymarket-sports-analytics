import os
import time
import sys
import argparse
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Optional

# Import from shared modules
from utils.shared_utils import (
    SPORTS_CONFIG,
    LATE_PICK_THRESHOLD,
    MIN_WIN_PCT,
    MIN_GAMES_FOR_WIN_PCT,
    MIN_GAMES,
    MIN_WINS,
    coerce_numeric_series,
    parse_game_teams,
    get_season,
    get_current_week,
    get_last_n_weeks,
    filter_by_season,
    filter_by_weeks,
    get_output_filename,
)
from utils.excel_utils import (
    GREEN_FILL,
    RED_FILL,
    YELLOW_FILL,
    HEADER_FILL,
    HEADER_FONT,
    format_time,
)
from utils.menu_utils import select_sport, select_season, select_time_period

WRITE_PROGRESS_THRESHOLD = 2000


# =============================================================================
# Data Transformation
# =============================================================================

def load_and_transform(
    input_file: str,
    late_pick_threshold: float = LATE_PICK_THRESHOLD,
    min_win_pct: float = MIN_WIN_PCT,
    min_games_for_win_pct: int = MIN_GAMES_FOR_WIN_PCT,
    min_games: int = MIN_GAMES,
    min_wins: int = MIN_WINS,
) -> pd.DataFrame:
    """
    Load trade data from Parquet and transform to expected format.

    Transformations:
        - match_title -> game
        - game_start_time -> game_date (extract date)
        - is_correct_pick -> result (True->won, False->lost, NA->pending)
    """
    df = pd.read_parquet(input_file)
    print(f"Load: {input_file} ({len(df):,} rows)")

    # Validate required columns exist
    required_cols = ['is_correct_pick', 'game_start_time', 'match_title',
                     'user_pick', 'yes_avg_price', 'no_avg_price', 'user_address']
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    # Extract game_date from game_start_time (native datetime from Parquet)
    df['game_date'] = df['game_start_time'].dt.strftime('%Y-%m-%d')

    # Create unique game identifier (match_title + date) to handle rematches
    df['game'] = df['match_title'] + ' (' + df['game_date'] + ')'

    # Map is_correct_pick to result (already nullable bool from Parquet)
    df['result'] = df['is_correct_pick'].map({True: 'won', False: 'lost'}).fillna('pending')

    # Calculate user's entry price for their pick
    # If user_pick matches first team (Team A), use yes_avg_price, else no_avg_price
    def get_pick_price(row):
        team_a, _ = parse_game_teams(row['match_title'])
        if row['user_pick'] == team_a:
            price = row.get('yes_avg_price')
            return price if price is not None and not pd.isna(price) else 0
        else:
            price = row.get('no_avg_price')
            return price if price is not None and not pd.isna(price) else 0

    df['pick_price'] = df.apply(get_pick_price, axis=1)

    # Filter out individual late picks
    original_picks = len(df)
    df = df[df['pick_price'] < late_pick_threshold]
    excluded_picks = original_picks - len(df)

    # Filter users by win rate (>= 70% with minimum 5 games)
    users_before_filters = df['user_address'].nunique()
    user_stats = df.groupby('user_address').agg(
        total_games=('result', lambda x: x.isin(['won', 'lost']).sum()),
        wins=('result', lambda x: (x == 'won').sum())
    ).reset_index()

    user_stats['win_pct'] = (user_stats['wins'] / user_stats['total_games'] * 100).fillna(0)

    # Only filter users with >= 5 games; keep users with < 5 games (not enough data)
    qualified_users = user_stats[
        (user_stats['total_games'] < min_games_for_win_pct) |  # Keep users with < min_games_for_win_pct
        (user_stats['win_pct'] >= min_win_pct)       # Keep users with >= min_win_pct win rate
    ]['user_address']

    original_users = df['user_address'].nunique()
    df = df[df['user_address'].isin(qualified_users)]
    excluded_win_rate_users = original_users - df['user_address'].nunique()

    # Filter users by minimum games (>= 3 games)
    user_game_counts = df.groupby('user_address').size()
    users_with_min_games = user_game_counts[user_game_counts >= min_games].index
    original_users = df['user_address'].nunique()
    df = df[df['user_address'].isin(users_with_min_games)]
    excluded_min_games_users = original_users - df['user_address'].nunique()

    # Filter users by minimum wins (>= 3 wins)
    user_wins = df[df['result'] == 'won'].groupby('user_address').size()
    users_with_min_wins = user_wins[user_wins >= min_wins].index
    original_users = df['user_address'].nunique()
    df = df[df['user_address'].isin(users_with_min_wins)]
    excluded_min_wins_users = original_users - df['user_address'].nunique()
    final_users = df['user_address'].nunique()

    print(
        "Filter: "
        f"picks {original_picks:,}->{len(df):,} (late removed {excluded_picks:,}); "
        f"users {users_before_filters:,}->{final_users:,} "
        f"(win-rate {excluded_win_rate_users:,}, min-games {excluded_min_games_users:,}, min-wins {excluded_min_wins_users:,})"
    )

    # Detect duplicate picks (same user, same game)
    duplicates = df.groupby(['user_address', 'game']).size()
    duplicate_count = (duplicates > 1).sum()
    if duplicate_count > 0:
        print(f"   Warning: {duplicate_count} user-game combinations have multiple picks (using first)")

    return df


# =============================================================================
# Excel Generation
# =============================================================================

def generate_excel(df: pd.DataFrame, output_file: str, title: str):
    """Generate Excel leaderboard from filtered DataFrame using vectorized operations."""

    if df.empty:
        print("No data to generate Excel from")
        return

    start_time = time.time()

    # Get unique games ordered by game_start_time
    sort_col = "game_start_time" if "game_start_time" in df.columns else "game_date"
    games_df = df[["game", sort_col]].drop_duplicates().sort_values(sort_col)
    games = games_df["game"].tolist()
    # Get unique users
    total_users = df["user_address"].nunique()

    # Group by user and compute stats in one pass
    stats_df = df.groupby("user_address").agg(
        wins=("result", lambda x: (x == "won").sum()),
        losses=("result", lambda x: (x == "lost").sum()),
        pending=("result", lambda x: (x == "pending").sum()),
    ).reset_index()

    # Calculate games (total picks made)
    stats_df["games"] = stats_df["wins"] + stats_df["losses"] + stats_df["pending"]

    # Calculate win percentage (keep NaN for users with no decided games)
    stats_df["total_decided"] = stats_df["wins"] + stats_df["losses"]
    stats_df["win_pct"] = (100 * stats_df["wins"] / stats_df["total_decided"]).round(1)
    # Don't fill NaN with 0 - users with no decided games should show blank, not 0%
    stats_df = stats_df.drop(columns=["total_decided"])

    # Optional PNL/ROI stats (if columns exist)
    if {"total_pnl", "yes_total_bought", "no_total_bought"}.issubset(df.columns):
        df['total_pnl_num'] = coerce_numeric_series(df['total_pnl']).fillna(0.0)
        df['total_bought_num'] = (
            coerce_numeric_series(df['yes_total_bought']).fillna(0.0) +
            coerce_numeric_series(df['no_total_bought']).fillna(0.0)
        )

        pnl_stats = df.groupby("user_address").agg(
            total_pnl=("total_pnl_num", "sum"),
            total_bought=("total_bought_num", "sum"),
        ).reset_index()
        pnl_stats["roi_pct"] = (
            100 * pnl_stats["total_pnl"] / pnl_stats["total_bought"]
        ).round(2)

        stats_df = stats_df.merge(pnl_stats, on="user_address", how="left")


    # Filter to resolved games only for streak calculation (exclude pending)
    df_resolved = df[df['result'].isin(['won', 'lost'])].copy()
    # Add original index as stable tiebreaker for deterministic ordering
    df_resolved['_orig_idx'] = range(len(df_resolved))
    # Sort by time descending, using original index as tiebreaker for same timestamps
    df_sorted = df_resolved.sort_values(
        [sort_col, '_orig_idx'],
        ascending=[False, False]  # Most recent first, original order for same-time games
    )

    # Calculate streaks using groupby
    def calc_streaks(group):
        results = group["result"].tolist()
        win_streak = 0
        loss_streak = 0

        # Win streak
        for r in results:
            if r == "won":
                win_streak += 1
            elif r == "lost":
                break

        # Loss streak
        for r in results:
            if r == "lost":
                loss_streak += 1
            elif r == "won":
                break

        return pd.Series({"win_streak": win_streak, "loss_streak": loss_streak})

    streaks_df = df_sorted.groupby("user_address").apply(calc_streaks, include_groups=False).reset_index()

    # Merge stats and streaks
    stats_df = stats_df.merge(streaks_df, on="user_address", how="left")
    stats_df["win_streak"] = stats_df["win_streak"].fillna(0).astype(int)
    stats_df["loss_streak"] = stats_df["loss_streak"].fillna(0).astype(int)


    # Only consider resolved games for last 10 calculation
    resolved_df = df[df['result'].isin(['won', 'lost'])].copy()

    if not resolved_df.empty:
        # Parse game_start_time with UTC to handle mixed timezones
        resolved_df['game_time_parsed'] = pd.to_datetime(resolved_df['game_start_time'], utc=True)

        # Sort by user and time descending, take last 10 per user
        resolved_df = resolved_df.sort_values(['user_address', 'game_time_parsed'], ascending=[True, False])
        last_10_df = resolved_df.groupby('user_address').head(10)

        # Calculate last 10 stats
        last_10_stats = last_10_df.groupby('user_address').agg(
            last_10_wins=('result', lambda x: (x == 'won').sum()),
            last_10_games=('result', 'count')
        ).reset_index()

        # Merge with stats_df
        stats_df = stats_df.merge(last_10_stats, on='user_address', how='left')
        stats_df['last_10_wins'] = stats_df['last_10_wins'].fillna(0).astype(int)
        stats_df['last_10_games'] = stats_df['last_10_games'].fillna(0).astype(int)
    else:
        stats_df['last_10_wins'] = 0
        stats_df['last_10_games'] = 0


    if not df.empty:
        # Determine majority pick per game
        majority_pick = (
            df.groupby("game")["user_pick"]
            .agg(lambda x: x.value_counts().index[0] if len(x) > 0 else "")
        )
        df = df.join(majority_pick.rename("majority_pick"), on="game")
        df["is_majority_pick"] = df["user_pick"] == df["majority_pick"]

        resolved_df = df[df["result"].isin(["won", "lost"])].copy()
        resolved_df["contrarian_game"] = (~resolved_df["is_majority_pick"]).astype(int)
        resolved_df["contrarian_win"] = (
            (~resolved_df["is_majority_pick"]) & (resolved_df["result"] == "won")
        ).astype(int)

        contrarian_stats = resolved_df.groupby("user_address").agg(
            contrarian_games=("contrarian_game", "sum"),
            contrarian_wins=("contrarian_win", "sum"),
        ).reset_index()
        contrarian_stats["contrarian_win_pct"] = (
            100 * contrarian_stats["contrarian_wins"] / contrarian_stats["contrarian_games"]
        ).round(1)

        stats_df = stats_df.merge(contrarian_stats, on="user_address", how="left")


    # Ensure optional columns exist for consistent output
    for col in ["roi_pct", "total_pnl", "contrarian_win_pct", "contrarian_games"]:
        if col not in stats_df.columns:
            stats_df[col] = None

    # Pivot picks
    picks_pivot = df.pivot_table(
        index="user_address",
        columns="game",
        values="user_pick",
        aggfunc="first"
    )

    # Pivot results
    results_pivot = df.pivot_table(
        index="user_address",
        columns="game",
        values="result",
        aggfunc="first"
    )

    # Rename result columns
    results_pivot.columns = [f"{col}_result" for col in results_pivot.columns]

    # Combine pivots
    pivot_df = picks_pivot.join(results_pivot).reset_index()

    # Build consensus data for each game (team names and date extracted separately)
    game_consensus = {}
    for game_name in games:
        # Game name format: "Team A vs Team B (YYYY-MM-DD)" or "Team A vs. Team B (YYYY-MM-DD)"
        # Extract date from end (last 12 chars are " (YYYY-MM-DD)")
        if game_name.endswith(')') and len(game_name) > 13:
            game_date = game_name[-11:-1]  # Extract "YYYY-MM-DD"
            match_title = game_name[:-13]   # Remove " (YYYY-MM-DD)"
        else:
            game_date = ""
            match_title = game_name

        # Parse team names from match title
        team_a, team_b = parse_game_teams(match_title)
        if not team_a:
            team_a = "Team A"
            team_b = "Team B"

        game_consensus[game_name] = {
            'team_a': team_a,
            'team_b': team_b,
            'match_title': match_title,
            'game_date': game_date,
        }

    result_df = stats_df.merge(pivot_df, on="user_address", how="left")

    # Sort by wins desc, win_pct desc, loss_streak asc
    result_df = result_df.sort_values(
        ["wins", "win_pct", "loss_streak"],
        ascending=[False, False, True]
    ).reset_index(drop=True)

    # Add rank column
    result_df.insert(0, "rank", range(1, len(result_df) + 1))

    # Fill NaN with empty string for display
    result_df = result_df.fillna("")

    print(f"Build leaderboard: {len(result_df):,} users, {len(games):,} games")
    print(f"Write workbook: {output_file}")

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Define columns to write (order: rank, user_address, games, wins, losses, win_pct,
    # win/loss streaks, last_10, ROI/PNL, contrarian stats, then game columns)
    stats_cols = [
        "rank",
        "user_address",
        "games",
        "wins",
        "losses",
        "win_pct",
        "win_streak",
        "loss_streak",
        "last_10",
        "roi_pct",
        "total_pnl",
        "contrarian_win_pct",
        "contrarian_games",
    ]
    display_cols = stats_cols + games
    num_stats_cols = len(stats_cols)

    # Create display column for last_10 (just the wins count)
    result_df["last_10"] = result_df["last_10_wins"].fillna(0).astype(int)

    # Header display names mapping
    header_names = {
        "rank": "rank",
        "user_address": "user_address",
        "games": "games",
        "wins": "wins",
        "losses": "losses",
        "win_pct": "win %",
        "win_streak": "win streak",
        "loss_streak": "loss streak",
        "last_10": "last 10",
        "roi_pct": "roi %",
        "total_pnl": "total pnl",
        "contrarian_win_pct": "contrarian win %",
        "contrarian_games": "contrarian games",
    }

    center_align = Alignment(horizontal="center")

    # Calculate total rows for data range
    total_rows = len(result_df)
    header_row_count = 6

    # ==========================================================================
    # Write 6 header rows with consensus data
    # Row 1: Team A consensus % (formula)
    # Row 2: Team A name (clean, no date)
    # Row 3: Team B consensus % (formula)
    # Row 4: Team B name (clean, no date)
    # Row 5: Game date
    # Row 6: Match title (main header, "Team A vs Team B" without date)
    # ==========================================================================

    # Calculate data range for formulas (data starts at row 7)
    data_start_row = header_row_count + 1
    data_end_row = data_start_row + total_rows - 1

    # Row 1: Team A consensus percentages (formulas)
    for col_idx, col_name in enumerate(display_cols, 1):
        if col_idx <= num_stats_cols:
            # Stats columns - empty for rows 1-4
            cell = ws.cell(row=1, column=col_idx, value="")
        else:
            # Game columns - Team A % formula
            consensus = game_consensus.get(col_name, {})
            team_a = consensus.get('team_a', '')
            # Escape double quotes for Excel formula
            team_a_escaped = team_a.replace('"', '""')
            col_letter = get_column_letter(col_idx)
            # Formula: COUNTIF(range, team_name) / COUNTA(range) with IFERROR to handle division by zero
            formula = f'=IFERROR(COUNTIF({col_letter}{data_start_row}:{col_letter}{data_end_row},"{team_a_escaped}")/COUNTA({col_letter}{data_start_row}:{col_letter}{data_end_row}),"")'
            cell = ws.cell(row=1, column=col_idx, value=formula)
            cell.number_format = '0%'
        cell.alignment = center_align
        cell.font = Font(bold=True)

    # Row 2: Team A names
    for col_idx, col_name in enumerate(display_cols, 1):
        if col_idx <= num_stats_cols:
            cell = ws.cell(row=2, column=col_idx, value="")
        else:
            consensus = game_consensus.get(col_name, {})
            cell = ws.cell(row=2, column=col_idx, value=consensus.get('team_a', ''))
        cell.alignment = center_align
        cell.font = Font(bold=True)

    # Row 3: Team B consensus percentages (formulas)
    for col_idx, col_name in enumerate(display_cols, 1):
        if col_idx <= num_stats_cols:
            cell = ws.cell(row=3, column=col_idx, value="")
        else:
            consensus = game_consensus.get(col_name, {})
            team_b = consensus.get('team_b', '')
            # Escape double quotes for Excel formula
            team_b_escaped = team_b.replace('"', '""')
            col_letter = get_column_letter(col_idx)
            # Formula: COUNTIF(range, team_name) / COUNTA(range) with IFERROR to handle division by zero
            formula = f'=IFERROR(COUNTIF({col_letter}{data_start_row}:{col_letter}{data_end_row},"{team_b_escaped}")/COUNTA({col_letter}{data_start_row}:{col_letter}{data_end_row}),"")'
            cell = ws.cell(row=3, column=col_idx, value=formula)
            cell.number_format = '0%'
        cell.alignment = center_align
        cell.font = Font(bold=True)

    # Row 4: Team B names
    for col_idx, col_name in enumerate(display_cols, 1):
        if col_idx <= num_stats_cols:
            cell = ws.cell(row=4, column=col_idx, value="")
        else:
            consensus = game_consensus.get(col_name, {})
            cell = ws.cell(row=4, column=col_idx, value=consensus.get('team_b', ''))
        cell.alignment = center_align
        cell.font = Font(bold=True)

    # Row 5: Game dates
    for col_idx, col_name in enumerate(display_cols, 1):
        if col_idx <= num_stats_cols:
            cell = ws.cell(row=5, column=col_idx, value="")
        else:
            consensus = game_consensus.get(col_name, {})
            cell = ws.cell(row=5, column=col_idx, value=consensus.get('game_date', ''))
        cell.alignment = center_align
        cell.font = Font(bold=True)

    # Row 6: Main header row (stats headers + match titles without date)
    for col_idx, col_name in enumerate(display_cols, 1):
        if col_idx <= num_stats_cols:
            display_name = header_names.get(col_name, col_name)
        else:
            # For game columns, use clean match_title without date
            consensus = game_consensus.get(col_name, {})
            display_name = consensus.get('match_title', col_name)
        cell = ws.cell(row=6, column=col_idx, value=display_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = center_align

    # Convert games to set for fast lookup
    games_set = set(games)

    # Write data rows (starting at row 7)
    show_progress = total_rows >= WRITE_PROGRESS_THRESHOLD
    for row_idx in range(total_rows):
        if show_progress and (row_idx + 1) % 10000 == 0:
            pct = ((row_idx + 1) / total_rows) * 100
            print(f"   Writing row {row_idx + 1:,}/{total_rows:,} ({pct:.1f}%)", end="\r")

        row_dict = result_df.iloc[row_idx]
        excel_row = row_idx + header_row_count + 1  # +1 for 1-indexed, +header_row_count for headers

        for col_idx, col_name in enumerate(display_cols, 1):
            value = row_dict.get(col_name, "")
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.alignment = center_align

            # Add hyperlink for user_address column
            if col_name == "user_address" and value:
                cell.hyperlink = f"https://polymarket.com/profile/{value}"
                cell.font = Font(color="0563C1", underline="single")

            # Number formats for stats
            if col_name in {"win_pct", "roi_pct", "contrarian_win_pct"} and value != "":
                cell.number_format = "0.0"
            if col_name == "total_pnl" and value != "":
                cell.number_format = "$#,##0.00"

            # Apply color formatting for game columns
            if col_name in games_set:
                result_col = f"{col_name}_result"
                result = row_dict.get(result_col, "")

                if result == "won":
                    cell.fill = GREEN_FILL
                elif result == "lost":
                    cell.fill = RED_FILL
                elif result == "pending":
                    cell.fill = YELLOW_FILL

    if show_progress:
        print(f"   Writing row {total_rows:,}/{total_rows:,} (100.0%)")

    # Add freeze pane after stats columns and header rows
    freeze_col = get_column_letter(num_stats_cols + 1)
    ws.freeze_panes = f"{freeze_col}{header_row_count + 1}"

    # Set column widths
    stats_widths = {
        "rank": 6,
        "user_address": 20,
        "games": 7,
        "wins": 6,
        "losses": 7,
        "win_pct": 7,
        "win_streak": 10,
        "loss_streak": 10,
        "last_10": 7,
        "roi_pct": 7,
        "total_pnl": 10,
        "contrarian_win_pct": 14,
        "contrarian_games": 14,
    }
    for col_idx, col_name in enumerate(stats_cols, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = stats_widths.get(col_name, 10)

    # Game columns get minimum width of 10
    for col_idx in range(num_stats_cols + 1, len(display_cols) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 10

    # Save
    wb.save(output_file)

    total_time = time.time() - start_time
    print(f"Done: {len(result_df):,} users, {total_rows:,} workbook rows -> {output_file} ({format_time(total_time)})")


# =============================================================================
# Main Operations
# =============================================================================

def do_generate(
    sport: str,
    weeks: Optional[List[int]] = None,
    is_season: bool = False,
    season_id: Optional[str] = None,
    late_pick_threshold: float = LATE_PICK_THRESHOLD,
    min_win_pct: float = MIN_WIN_PCT,
    min_games_for_win_pct: int = MIN_GAMES_FOR_WIN_PCT,
    min_games: int = MIN_GAMES,
    min_wins: int = MIN_WINS,
):
    """Main generation function."""
    if sport.lower() == "all":
        input_file = "db_trades.parquet"
        season_label = str(datetime.now().year)
    else:
        config = SPORTS_CONFIG[sport]
        input_file = config["input_file"]
        season = get_season(sport, season_id)
        season_id = season["season_id"]
        season_label = season["label"]

    # Check if input file exists
    if not os.path.exists(input_file):
        print(f"Input file not found: {input_file}")
        return

    # Load and transform data
    df = load_and_transform(
        input_file,
        late_pick_threshold=late_pick_threshold,
        min_win_pct=min_win_pct,
        min_games_for_win_pct=min_games_for_win_pct,
        min_games=min_games,
        min_wins=min_wins,
    )

    # Always constrain single-sport reports to the selected season window
    if sport.lower() != "all":
        print(f"Filtering to {sport} {season_label} season window...")
        df = filter_by_season(df, sport, season_id)
        print(f"   Filtered to {len(df):,} rows")

    # Filter by weeks if specified
    if weeks and not is_season and sport.lower() != "all":
        week_range = f"{min(weeks)}-{max(weeks)}" if len(weeks) > 1 else str(weeks[0])
        print(f"Filtering to Week(s) {week_range}...")
        df = filter_by_weeks(df, weeks, sport, season_id)
        print(f"   Filtered to {len(df):,} rows")
        title = f"{sport} {season_label} Weeks {week_range} (Postseason Included)"
    else:
        if sport.lower() == "all":
            title = f"ALL Sports Season {season_label}"
        else:
            title = f"{sport} {season_label} Season (Postseason Included)"

    if df.empty:
        print("No data matches the filter criteria")
        return

    # Generate output filename
    output_file = get_output_filename(
        weeks,
        is_season,
        sport,
        prefix="leaderboard",
        season_id=season_id,
    )

    # Generate Excel
    generate_excel(df, output_file, title)


# =============================================================================
# Main Entry Point
# =============================================================================

def main():
    if len(sys.argv) > 1:
        parser = argparse.ArgumentParser(description="Generate leaderboard Excel")
        parser.add_argument(
            "--sport",
            type=str,
            required=True,
            help="Sport: NFL, NBA, CFB, CBB, or all",
        )
        parser.add_argument("--weeks", type=str, help="Weeks list (e.g., 5 or 3-5 or 1,2,3)")
        parser.add_argument("--season-id", type=str, help="Season ID (e.g., 2025, 2025-26)")
        parser.add_argument("--season", action="store_true", help="Full season")
        parser.add_argument("--latest", action="store_true", help="Latest week")
        parser.add_argument("--previous", action="store_true", help="Previous week")
        parser.add_argument("--last5", action="store_true", help="Last 5 weeks")
        parser.add_argument("--late-pick-threshold", type=float, default=LATE_PICK_THRESHOLD)
        parser.add_argument("--min-win-pct", type=float, default=MIN_WIN_PCT)
        parser.add_argument("--min-games-win-pct", type=int, default=MIN_GAMES_FOR_WIN_PCT)
        parser.add_argument("--min-games", type=int, default=MIN_GAMES)
        parser.add_argument("--min-wins", type=int, default=MIN_WINS)

        args = parser.parse_args()

        sport_arg = args.sport.strip().lower()
        sport = "all" if sport_arg == "all" else sport_arg.upper()
        selected_season_id = None

        if sport != "all":
            try:
                selected_season = get_season(sport, args.season_id)
                selected_season_id = selected_season["season_id"]
            except ValueError as e:
                print(f"Error: {e}")
                return
        elif args.season_id:
            print("Warning: --season-id is ignored when --sport all")

        def parse_weeks_arg(value: str) -> List[int]:
            value = value.strip()
            if "-" in value:
                start, end = value.split("-", 1)
                return list(range(int(start), int(end) + 1))
            if "," in value:
                return [int(v.strip()) for v in value.split(",") if v.strip()]
            return [int(value)]

        # Determine weeks/season
        weeks = None
        is_season = True

        if args.season:
            weeks = None
            is_season = True
        elif args.latest or args.previous or args.last5 or args.weeks:
            if sport == "all":
                print("Week filtering is not supported for all sports. Generating full season.")
                weeks = None
                is_season = True
            else:
                current_week = get_current_week(sport, selected_season_id, include_postseason=True)
                if args.latest:
                    if current_week == 0:
                        print("Season has not started yet. Generating full season.")
                        weeks = None
                    else:
                        weeks = [current_week]
                elif args.previous:
                    if current_week <= 1:
                        print("Previous week is not available. Generating latest available week.")
                        weeks = [current_week] if current_week > 0 else None
                    else:
                        weeks = [current_week - 1]
                elif args.last5:
                    weeks = get_last_n_weeks(
                        5, sport, selected_season_id, include_postseason=True
                    )
                    if not weeks:
                        print("No active week found yet. Generating full season.")
                        weeks = None
                elif args.weeks:
                    weeks = parse_weeks_arg(args.weeks)
                is_season = weeks is None

        do_generate(
            sport=sport,
            weeks=weeks,
            is_season=is_season,
            season_id=selected_season_id,
            late_pick_threshold=args.late_pick_threshold,
            min_win_pct=args.min_win_pct,
            min_games_for_win_pct=args.min_games_win_pct,
            min_games=args.min_games,
            min_wins=args.min_wins,
        )
        return

    # Interactive mode
    sport = select_sport(title="Leaderboard Generator", include_all=True)
    if not sport or sport == "exit":
        return

    if sport == "all":
        do_generate(sport=sport, weeks=None, is_season=True)
        return

    season_id = select_season(sport)
    if not season_id or season_id == "exit":
        return

    weeks, is_season = select_time_period(sport, season_id)
    if weeks is None and not is_season:
        return

    do_generate(sport=sport, weeks=weeks, is_season=is_season, season_id=season_id)


if __name__ == "__main__":
    main()
