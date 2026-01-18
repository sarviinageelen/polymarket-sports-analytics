"""
Excel utilities for Polymarket Sports Analytics.

This module contains Excel formatting constants, cell styling helpers,
and progress display utilities shared across report generators.
"""

from openpyxl.styles import Font, PatternFill, Alignment


# =============================================================================
# Color Constants
# =============================================================================

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


# =============================================================================
# Cell Styling Helpers
# =============================================================================

def apply_result_color(cell, result: str) -> None:
    """
    Apply color fill to cell based on result.

    Args:
        cell: openpyxl cell object
        result: Result string ("won", "lost", or "pending")
    """
    if result == "won":
        cell.fill = GREEN_FILL
    elif result == "lost":
        cell.fill = RED_FILL
    elif result == "pending":
        cell.fill = YELLOW_FILL


def apply_header_style(cell) -> None:
    """
    Apply header styling to cell (blue fill, white bold font).

    Args:
        cell: openpyxl cell object
    """
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT


def create_hyperlink_cell(cell, url: str, display_text: str = None) -> None:
    """
    Create a clickable hyperlink cell.

    Args:
        cell: openpyxl cell object
        url: URL to link to
        display_text: Optional text to display (defaults to cell value)
    """
    if display_text:
        cell.value = display_text
    cell.hyperlink = url
    cell.font = Font(color="0563C1", underline="single")


# =============================================================================
# Progress Display Utilities
# =============================================================================

def progress_bar(current: int, total: int, width: int = 40) -> str:
    """
    Generate a progress bar string.

    Args:
        current: Current progress count
        total: Total items to process
        width: Width of progress bar in characters

    Returns:
        Formatted progress bar string like "[====----] 50/100 (50.0%)"
    """
    if total == 0:
        return f"[{'=' * width}] 0/0"

    pct = current / total
    filled = int(width * pct)
    bar = '=' * filled + '-' * (width - filled)
    return f"[{bar}] {current}/{total} ({pct*100:.1f}%)"


def print_progress(current: int, total: int, prefix: str = "", suffix: str = "") -> None:
    """
    Print progress bar that updates in place.

    Args:
        current: Current progress count
        total: Total items to process
        prefix: Text to prepend to progress bar
        suffix: Text to append to progress bar
    """
    bar = progress_bar(current, total)
    end_char = '\n' if current >= total else '\r'
    line = f"{prefix}{bar} {suffix}".ljust(80)
    print(line, end=end_char, flush=True)


def format_time(seconds: float) -> str:
    """
    Format seconds into human-readable time.

    Args:
        seconds: Time in seconds

    Returns:
        Formatted string like "1h 23m" or "45 sec"
    """
    if seconds < 60:
        return f"{int(seconds)} sec"
    elif seconds < 3600:
        minutes = int(seconds / 60)
        secs = int(seconds % 60)
        return f"{minutes}m {secs}s"
    else:
        hours = int(seconds / 3600)
        minutes = int((seconds % 3600) / 60)
        return f"{hours}h {minutes}m"


def calculate_eta(elapsed: float, current: int, total: int) -> str:
    """
    Calculate estimated time remaining.

    Args:
        elapsed: Elapsed time in seconds
        current: Current progress (items completed)
        total: Total items to process

    Returns:
        Formatted ETA string
    """
    if current == 0:
        return "calculating..."

    avg_time_per_item = elapsed / current
    remaining_items = total - current
    eta_seconds = avg_time_per_item * remaining_items

    return format_time(eta_seconds)
