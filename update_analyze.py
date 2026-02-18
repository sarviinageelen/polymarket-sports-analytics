import os
import time
import sys
import argparse
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Optional

# Import from shared modules
from utils.shared_utils import (
    SPORTS_CONFIG,
    LATE_PICK_THRESHOLD,
    MIN_WIN_PCT,
    MIN_GAMES_FOR_WIN_PCT,
    MIN_GAMES,
    MIN_WINS,
    coerce_numeric_series,
    parse_game_teams,
    get_season,
    get_current_week,
    get_last_n_weeks,
    filter_by_season,
    filter_by_weeks,
    get_output_filename,
)
from utils.excel_utils import (
    GREEN_FILL,
    RED_FILL,
    YELLOW_FILL,
    HEADER_FILL,
    HEADER_FONT,
    format_time,
)
from utils.menu_utils import select_sport, select_season, select_time_period

WRITE_PROGRESS_THRESHOLD = 2000


# =============================================================================
# Data Loading and Transformation
# =============================================================================

def load_and_transform(
    input_file: str,
    late_pick_threshold: float = LATE_PICK_THRESHOLD,
    min_win_pct: float = MIN_WIN_PCT,
    min_games_for_win_pct: int = MIN_GAMES_FOR_WIN_PCT,
    min_games: int = MIN_GAMES,
    min_wins: int = MIN_WINS,
) -> pd.DataFrame:
    """
    Load trade data from Parquet and transform to expected format.

    Similar to update_picks.py but with additional filters for min games/wins.
    """
    df = pd.read_parquet(input_file)
    print(f"Load: {input_file} ({len(df):,} rows)")

    # Extract game_date from game_start_time (native datetime from Parquet)
    df['game_date'] = df['game_start_time'].dt.strftime('%Y-%m-%d')

    # Create unique game identifier
    df['game'] = df['match_title'] + ' (' + df['game_date'] + ')'

    # Map is_correct_pick to result (already nullable bool from Parquet)
    df['result'] = df['is_correct_pick'].map({True: 'won', False: 'lost'}).fillna('pending')

    # Coerce price columns to numeric (safe no-op for Parquet data)
    df['yes_avg_price'] = coerce_numeric_series(df.get('yes_avg_price'))
    df['no_avg_price'] = coerce_numeric_series(df.get('no_avg_price'))

    def get_pick_price(row):
        team_a, _ = parse_game_teams(row['game'])
        if row['user_pick'] == team_a:
            price = row.get('yes_avg_price', 0)
            return price if price is not None and not pd.isna(price) else 0
        else:
            price = row.get('no_avg_price', 0)
            return price if price is not None and not pd.isna(price) else 0

    df['pick_price'] = df.apply(get_pick_price, axis=1)

    original_picks = len(df)
    df = df[df['pick_price'] < late_pick_threshold]
    excluded_picks = original_picks - len(df)

    # Filter users by win rate (>= 70% with minimum 5 games)
    users_before_filters = df['user_address'].nunique()
    user_stats = df.groupby('user_address').agg(
        total_games=('result', lambda x: x.isin(['won', 'lost']).sum()),
        wins=('result', lambda x: (x == 'won').sum())
    ).reset_index()

    user_stats['win_pct'] = (user_stats['wins'] / user_stats['total_games'] * 100).fillna(0)

    qualified_users = user_stats[
        (user_stats['total_games'] < min_games_for_win_pct) |
        (user_stats['win_pct'] >= min_win_pct)
    ]['user_address']

    original_users = df['user_address'].nunique()
    df = df[df['user_address'].isin(qualified_users)]
    excluded_win_rate_users = original_users - df['user_address'].nunique()

    # Filter users by minimum games
    user_game_counts = df.groupby('user_address').size()
    users_with_min_games = user_game_counts[user_game_counts >= min_games].index
    original_users = df['user_address'].nunique()
    df = df[df['user_address'].isin(users_with_min_games)]
    excluded_min_games_users = original_users - df['user_address'].nunique()

    # Filter users by minimum wins
    user_wins = df[df['result'] == 'won'].groupby('user_address').size()
    users_with_min_wins = user_wins[user_wins >= min_wins].index
    original_users = df['user_address'].nunique()
    df = df[df['user_address'].isin(users_with_min_wins)]
    excluded_min_wins_users = original_users - df['user_address'].nunique()
    final_users = df['user_address'].nunique()

    print(
        "Filter: "
        f"picks {original_picks:,}->{len(df):,} (late removed {excluded_picks:,}); "
        f"users {users_before_filters:,}->{final_users:,} "
        f"(win-rate {excluded_win_rate_users:,}, min-games {excluded_min_games_users:,}, min-wins {excluded_min_wins_users:,})"
    )

    return df


# =============================================================================
# Stats Calculation Functions
# =============================================================================

def calculate_user_stats(df: pd.DataFrame) -> pd.DataFrame:
    """Calculate per-user statistics."""
    # Basic stats
    stats_df = df.groupby("user_address").agg(
        games=("result", "count"),
        wins=("result", lambda x: (x == "won").sum()),
        losses=("result", lambda x: (x == "lost").sum()),
    ).reset_index()

    # Win percentage
    stats_df["total_decided"] = stats_df["wins"] + stats_df["losses"]
    stats_df["win_pct"] = (100 * stats_df["wins"] / stats_df["total_decided"]).round(1)
    stats_df["win_pct"] = stats_df["win_pct"].fillna(0)
    stats_df = stats_df.drop(columns=["total_decided"])

    # Optional PNL/ROI stats
    if {"total_pnl", "yes_total_bought", "no_total_bought"}.issubset(df.columns):
        df['total_pnl_num'] = coerce_numeric_series(df['total_pnl']).fillna(0.0)
        df['total_bought_num'] = (
            coerce_numeric_series(df['yes_total_bought']).fillna(0.0) +
            coerce_numeric_series(df['no_total_bought']).fillna(0.0)
        )

        pnl_stats = df.groupby("user_address").agg(
            total_pnl=("total_pnl_num", "sum"),
            total_bought=("total_bought_num", "sum"),
        ).reset_index()
        pnl_stats["roi_pct"] = (
            100 * pnl_stats["total_pnl"] / pnl_stats["total_bought"]
        ).round(2)

        stats_df = stats_df.merge(pnl_stats, on="user_address", how="left")

    # Win streak calculation
    sort_col = "game_start_time" if "game_start_time" in df.columns else "game_date"
    df_resolved = df[df['result'].isin(['won', 'lost'])].copy()
    df_sorted = df_resolved.sort_values(
        [sort_col, 'is_correct_pick'],
        ascending=[False, True]
    )

    def calc_win_streak(group):
        results = group["result"].tolist()
        streak = 0
        for r in results:
            if r == "won":
                streak += 1
            elif r == "lost":
                break
        return streak

    streaks = df_sorted.groupby("user_address").apply(calc_win_streak, include_groups=False)
    streaks_df = streaks.reset_index()
    streaks_df.columns = ["user_address", "streak"]

    stats_df = stats_df.merge(streaks_df, on="user_address", how="left")
    stats_df["streak"] = stats_df["streak"].fillna(0).astype(int)

    # Last 10 calculation
    resolved_df = df[df['result'].isin(['won', 'lost'])].copy()

    if not resolved_df.empty:
        resolved_df['game_time_parsed'] = pd.to_datetime(resolved_df['game_start_time'], utc=True)
        resolved_df = resolved_df.sort_values(['user_address', 'game_time_parsed'], ascending=[True, False])
        last_10_df = resolved_df.groupby('user_address').head(10)

        last_10_stats = last_10_df.groupby('user_address').agg(
            last_10=('result', lambda x: (x == 'won').sum())
        ).reset_index()

        stats_df = stats_df.merge(last_10_stats, on='user_address', how='left')
        stats_df['last_10'] = stats_df['last_10'].fillna(0).astype(int)
    else:
        stats_df['last_10'] = 0

    # Ensure optional columns exist
    for col in ["total_pnl", "total_bought", "roi_pct"]:
        if col not in stats_df.columns:
            stats_df[col] = 0

    # Sort by win_pct for consistent ordering
    stats_df = stats_df.sort_values(
        ["win_pct", "wins", "streak"],
        ascending=[False, False, False]
    ).reset_index(drop=True)

    return stats_df


def calculate_consensus(df: pd.DataFrame) -> pd.DataFrame:
    """Calculate consensus percentages per game."""
    consensus_data = []

    for game_name in df['game'].unique():
        game_df = df[df['game'] == game_name]

        # Parse team names from game (format: "Team A vs Team B (YYYY-MM-DD)")
        if game_name.endswith(')') and len(game_name) > 13:
            match_title = game_name[:-13]
        else:
            match_title = game_name

        team_a, team_b = parse_game_teams(match_title)
        if not team_a:
            team_a = "Team A"
            team_b = "Team B"

        # Count picks
        total_picks = len(game_df)
        team_a_picks = (game_df['user_pick'] == team_a).sum()
        team_b_picks = (game_df['user_pick'] == team_b).sum()

        # Calculate percentages
        consensus_a = round(100 * team_a_picks / total_picks, 1) if total_picks > 0 else 0
        consensus_b = round(100 * team_b_picks / total_picks, 1) if total_picks > 0 else 0

        # Determine majority pick
        majority_pick = team_a if team_a_picks >= team_b_picks else team_b

        # Determine winner
        resolved = game_df[game_df['result'].isin(['won', 'lost'])]
        if len(resolved) > 0:
            won_row = resolved[resolved['result'] == 'won'].iloc[0] if len(resolved[resolved['result'] == 'won']) > 0 else None
            if won_row is not None:
                winner = won_row['user_pick']
            else:
                # All resolved picks lost, so winner is the other team
                lost_pick = resolved.iloc[0]['user_pick']
                winner = team_b if lost_pick == team_a else team_a
        else:
            winner = "Pending"

        consensus_data.append({
            'game': game_name,
            'team_a': team_a,
            'team_b': team_b,
            'consensus_a': consensus_a,
            'consensus_b': consensus_b,
            'majority_pick': majority_pick,
            'winner': winner,
        })

    consensus_df = pd.DataFrame(consensus_data)
    return consensus_df


# =============================================================================
# Build Flat Table
# =============================================================================

def build_flat_table(df: pd.DataFrame, user_stats: pd.DataFrame, consensus: pd.DataFrame) -> pd.DataFrame:
    """Build the flat table with one row per pick."""
    # Start with picks dataframe
    flat_df = df[['user_address', 'game', 'game_date', 'user_pick', 'result', 'pick_price']].copy()
    flat_df = flat_df.rename(columns={'user_pick': 'pick'})

    # Merge user stats
    flat_df = flat_df.merge(
        user_stats[['user_address', 'games', 'wins', 'losses', 'win_pct', 'streak', 'last_10', 'total_pnl', 'roi_pct']],
        on='user_address',
        how='left'
    )

    # Merge consensus
    flat_df = flat_df.merge(
        consensus[['game', 'team_a', 'team_b', 'consensus_a', 'consensus_b', 'majority_pick', 'winner']],
        on='game',
        how='left'
    )

    # Majority/contrarian flags
    flat_df['is_majority_pick'] = flat_df['pick'] == flat_df['majority_pick']
    flat_df['is_contrarian'] = ~flat_df['is_majority_pick']

    # Calculate pick_pct (% who picked same team as this user)
    flat_df['pick_pct'] = flat_df.apply(
        lambda row: row['consensus_a'] / 100 if row['pick'] == row['team_a'] else row['consensus_b'] / 100,
        axis=1
    )

    # Clean up game name (remove date suffix for cleaner display)
    flat_df['game_display'] = flat_df['game'].apply(
        lambda x: x[:-13] if x.endswith(')') and len(x) > 13 else x
    )

    # Convert game_date to datetime for Excel
    flat_df['game_date'] = pd.to_datetime(flat_df['game_date'])

    # Select and order columns
    flat_df = flat_df[[
        'user_address', 'games', 'wins', 'losses', 'win_pct', 'total_pnl', 'roi_pct', 'streak', 'last_10',
        'game_display', 'game_date', 'pick', 'result', 'pick_price', 'pick_pct',
        'majority_pick', 'is_majority_pick', 'is_contrarian', 'winner'
    ]]

    flat_df = flat_df.rename(columns={'game_display': 'game'})

    # Sort by win_pct desc, then game_date
    flat_df = flat_df.sort_values(['win_pct', 'game_date'], ascending=[False, True]).reset_index(drop=True)

    return flat_df


def build_market_summary(df: pd.DataFrame, consensus: pd.DataFrame) -> pd.DataFrame:
    """Build per-game summary table."""
    if df.empty:
        return pd.DataFrame()

    game_dates = df[['game', 'game_date']].drop_duplicates()
    total_counts = df.groupby('game').size().reset_index(name='total_picks')
    pick_counts = df.groupby(['game', 'user_pick']).size().reset_index(name='pick_count')

    summary = consensus.merge(game_dates, on='game', how='left')
    summary = summary.merge(total_counts, on='game', how='left')
    summary = summary.merge(
        pick_counts,
        left_on=['game', 'winner'],
        right_on=['game', 'user_pick'],
        how='left'
    )

    summary['winner_pick_pct'] = (summary['pick_count'] / summary['total_picks']).fillna(0.0)
    summary['majority_correct'] = summary['winner'] == summary['majority_pick']
    summary['consensus_a'] = (summary['consensus_a'] / 100).fillna(0.0)
    summary['consensus_b'] = (summary['consensus_b'] / 100).fillna(0.0)
    summary['game_date'] = pd.to_datetime(summary['game_date'], errors='coerce')

    summary = summary[[
        'game', 'game_date', 'total_picks',
        'team_a', 'team_b', 'consensus_a', 'consensus_b',
        'majority_pick', 'winner', 'winner_pick_pct', 'majority_correct'
    ]]

    return summary


# =============================================================================
# Excel Generation
# =============================================================================

def generate_excel_flat(flat_df: pd.DataFrame, output_file: str, title: str, market_summary: Optional[pd.DataFrame] = None):
    """Generate Excel file with flat table format."""
    if flat_df.empty:
        print("No data to generate Excel from")
        return

    start_time = time.time()
    print(f"Write workbook: {output_file}")

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Column headers
    columns = [
        ('user_address', 'User'),
        ('games', 'Games'),
        ('wins', 'Wins'),
        ('losses', 'Losses'),
        ('win_pct', 'Win %'),
        ('total_pnl', 'Total PnL'),
        ('roi_pct', 'ROI %'),
        ('streak', 'Streak'),
        ('last_10', 'Last 10'),
        ('game', 'Game'),
        ('game_date', 'Game Date'),
        ('pick', 'Pick'),
        ('result', 'Result'),
        ('pick_price', 'Price'),
        ('pick_pct', 'Pick %'),
        ('majority_pick', 'Majority Pick'),
        ('is_majority_pick', 'Majority?'),
        ('is_contrarian', 'Contrarian?'),
        ('winner', 'Winner'),
    ]

    center_align = Alignment(horizontal="center")

    # Write header row
    for col_idx, (col_key, col_name) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = center_align

    # Write data rows
    total_rows = len(flat_df)
    show_progress = total_rows >= WRITE_PROGRESS_THRESHOLD
    for row_idx, row in flat_df.iterrows():
        if show_progress and (row_idx + 1) % 5000 == 0:
            pct = ((row_idx + 1) / total_rows) * 100
            print(f"   Writing row {row_idx + 1:,}/{total_rows:,} ({pct:.1f}%)", end="\r")

        excel_row = row_idx + 2  # +1 for header, +1 for 1-indexed

        for col_idx, (col_key, col_name) in enumerate(columns, 1):
            value = row.get(col_key, "")

            # Handle game_date as Excel date
            if col_key == 'game_date' and pd.notna(value):
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.number_format = 'YYYY-MM-DD'
                cell.alignment = center_align
                continue

            # Handle percent fields
            if col_key in {'pick_pct'} and pd.notna(value):
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.number_format = '0.0%'
                cell.alignment = center_align
                continue

            if col_key in {'win_pct', 'roi_pct'} and pd.notna(value):
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.number_format = '0.0'
                cell.alignment = center_align
                continue

            if col_key == 'total_pnl' and pd.notna(value):
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.number_format = '$#,##0.00'
                cell.alignment = center_align
                continue

            # Format pick_price
            if col_key == 'pick_price' and pd.notna(value):
                value = round(value, 2)

            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.alignment = center_align

            # Add hyperlink for user_address column
            if col_key == "user_address" and value:
                cell.hyperlink = f"https://polymarket.com/profile/{value}"
                cell.font = Font(color="0563C1", underline="single")

            # Color code the pick column based on result
            if col_key == "pick":
                result = row.get('result', '')
                if result == "won":
                    cell.fill = GREEN_FILL
                elif result == "lost":
                    cell.fill = RED_FILL
                elif result == "pending":
                    cell.fill = YELLOW_FILL

    if show_progress:
        print(f"   Writing row {total_rows:,}/{total_rows:,} (100.0%)")

    # Freeze panes after user stats (after last_10)
    ws.freeze_panes = "J2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{total_rows + 1}"

    # Adjust column widths
    column_widths = {
        'A': 20,  # User
        'B': 7,   # Games
        'C': 6,   # Wins
        'D': 7,   # Losses
        'E': 7,   # Win %
        'F': 10,  # Total PnL
        'G': 7,   # ROI %
        'H': 7,   # Streak
        'I': 8,   # Last 10
        'J': 25,  # Game
        'K': 12,  # Game Date
        'L': 15,  # Pick
        'M': 8,   # Result
        'N': 7,   # Price
        'O': 8,   # Pick %
        'P': 15,  # Majority Pick
        'Q': 10,  # Majority?
        'R': 12,  # Contrarian?
        'S': 12,  # Winner
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Optional market summary sheet
    if market_summary is not None and not market_summary.empty:
        ws2 = wb.create_sheet(title="markets")

        summary_cols = [
            ('game', 'Game'),
            ('game_date', 'Game Date'),
            ('total_picks', 'Total Picks'),
            ('team_a', 'Team A'),
            ('team_b', 'Team B'),
            ('consensus_a', 'Consensus A %'),
            ('consensus_b', 'Consensus B %'),
            ('majority_pick', 'Majority Pick'),
            ('winner', 'Winner'),
            ('winner_pick_pct', 'Winner Pick %'),
            ('majority_correct', 'Majority Correct'),
        ]

        for col_idx, (col_key, col_name) in enumerate(summary_cols, 1):
            cell = ws2.cell(row=1, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = center_align

        for row_idx, row in market_summary.iterrows():
            excel_row = row_idx + 2
            for col_idx, (col_key, _) in enumerate(summary_cols, 1):
                value = row.get(col_key, "")
                cell = ws2.cell(row=excel_row, column=col_idx, value=value)
                cell.alignment = center_align

                if col_key in {'consensus_a', 'consensus_b', 'winner_pick_pct'} and pd.notna(value):
                    cell.number_format = '0.0%'
                if col_key == 'game_date' and pd.notna(value):
                    cell.number_format = 'YYYY-MM-DD'

        # Column widths for summary
        summary_widths = {
            'A': 30,
            'B': 12,
            'C': 12,
            'D': 20,
            'E': 20,
            'F': 14,
            'G': 14,
            'H': 18,
            'I': 14,
            'J': 14,
            'K': 16,
        }
        for col_letter, width in summary_widths.items():
            ws2.column_dimensions[col_letter].width = width

        ws2.freeze_panes = "A2"

    # Save
    wb.save(output_file)

    total_time = time.time() - start_time
    print(f"Done: {len(flat_df):,} rows -> {output_file} ({format_time(total_time)})")


# =============================================================================
# Main Operations
# =============================================================================

def do_generate(
    sport: str,
    weeks: Optional[List[int]] = None,
    is_season: bool = False,
    season_id: Optional[str] = None,
    late_pick_threshold: float = LATE_PICK_THRESHOLD,
    min_win_pct: float = MIN_WIN_PCT,
    min_games_for_win_pct: int = MIN_GAMES_FOR_WIN_PCT,
    min_games: int = MIN_GAMES,
    min_wins: int = MIN_WINS,
):
    """Main generation function."""
    if sport.lower() == "all":
        input_file = "db_trades.parquet"
        season_label = str(datetime.now().year)
    else:
        config = SPORTS_CONFIG[sport]
        input_file = config["input_file"]
        season = get_season(sport, season_id)
        season_id = season["season_id"]
        season_label = season["label"]

    if not os.path.exists(input_file):
        print(f"Input file not found: {input_file}")
        return

    # Load and transform data
    df = load_and_transform(
        input_file,
        late_pick_threshold=late_pick_threshold,
        min_win_pct=min_win_pct,
        min_games_for_win_pct=min_games_for_win_pct,
        min_games=min_games,
        min_wins=min_wins,
    )

    # Always constrain single-sport reports to the selected season window
    if sport.lower() != "all":
        print(f"Filtering to {sport} {season_label} season window...")
        df = filter_by_season(df, sport, season_id)
        print(f"   Filtered to {len(df):,} rows")

    # Filter by weeks if specified
    if weeks and not is_season and sport.lower() != "all":
        week_range = f"{min(weeks)}-{max(weeks)}" if len(weeks) > 1 else str(weeks[0])
        print(f"Filtering to Week(s) {week_range}...")
        df = filter_by_weeks(df, weeks, sport, season_id)
        print(f"   Filtered to {len(df):,} rows")
        title = f"{sport} {season_label} Weeks {week_range} (Postseason Included)"
    else:
        if sport.lower() == "all":
            title = f"ALL Sports Season {season_label}"
        else:
            title = f"{sport} {season_label} Season (Postseason Included)"

    if df.empty:
        print("No data matches the filter criteria")
        return

    # Calculate stats and consensus
    print("Build analysis table...")
    user_stats = calculate_user_stats(df)
    consensus = calculate_consensus(df)

    # Build flat table
    flat_df = build_flat_table(df, user_stats, consensus)
    market_summary = build_market_summary(df, consensus)

    # Generate output filename (using "analysis" prefix)
    output_file = get_output_filename(
        weeks,
        is_season,
        sport,
        prefix="analysis",
        season_id=season_id,
    )

    # Generate Excel
    generate_excel_flat(flat_df, output_file, title, market_summary=market_summary)


# =============================================================================
# Main Entry Point
# =============================================================================

def main():
    if len(sys.argv) > 1:
        parser = argparse.ArgumentParser(description="Generate flat-table analysis Excel")
        parser.add_argument(
            "--sport",
            type=str,
            required=True,
            help="Sport: NFL, NBA, CFB, CBB, or all",
        )
        parser.add_argument("--weeks", type=str, help="Weeks list (e.g., 5 or 3-5 or 1,2,3)")
        parser.add_argument("--season-id", type=str, help="Season ID (e.g., 2025, 2025-26)")
        parser.add_argument("--season", action="store_true", help="Full season")
        parser.add_argument("--latest", action="store_true", help="Latest week")
        parser.add_argument("--previous", action="store_true", help="Previous week")
        parser.add_argument("--last5", action="store_true", help="Last 5 weeks")
        parser.add_argument("--late-pick-threshold", type=float, default=LATE_PICK_THRESHOLD)
        parser.add_argument("--min-win-pct", type=float, default=MIN_WIN_PCT)
        parser.add_argument("--min-games-win-pct", type=int, default=MIN_GAMES_FOR_WIN_PCT)
        parser.add_argument("--min-games", type=int, default=MIN_GAMES)
        parser.add_argument("--min-wins", type=int, default=MIN_WINS)

        args = parser.parse_args()

        sport_arg = args.sport.strip().lower()
        sport = "all" if sport_arg == "all" else sport_arg.upper()
        selected_season_id = None

        if sport != "all":
            try:
                selected_season = get_season(sport, args.season_id)
                selected_season_id = selected_season["season_id"]
            except ValueError as e:
                print(f"Error: {e}")
                return
        elif args.season_id:
            print("Warning: --season-id is ignored when --sport all")

        def parse_weeks_arg(value: str) -> List[int]:
            value = value.strip()
            if "-" in value:
                start, end = value.split("-", 1)
                return list(range(int(start), int(end) + 1))
            if "," in value:
                return [int(v.strip()) for v in value.split(",") if v.strip()]
            return [int(value)]

        weeks = None
        is_season = True

        if args.season:
            weeks = None
            is_season = True
        elif args.latest or args.previous or args.last5 or args.weeks:
            if sport == "all":
                print("Week filtering is not supported for all sports. Generating full season.")
                weeks = None
                is_season = True
            else:
                current_week = get_current_week(sport, selected_season_id, include_postseason=True)
                if args.latest:
                    if current_week == 0:
                        print("Season has not started yet. Generating full season.")
                        weeks = None
                    else:
                        weeks = [current_week]
                elif args.previous:
                    if current_week <= 1:
                        print("Previous week is not available. Generating latest available week.")
                        weeks = [current_week] if current_week > 0 else None
                    else:
                        weeks = [current_week - 1]
                elif args.last5:
                    weeks = get_last_n_weeks(
                        5, sport, selected_season_id, include_postseason=True
                    )
                    if not weeks:
                        print("No active week found yet. Generating full season.")
                        weeks = None
                elif args.weeks:
                    weeks = parse_weeks_arg(args.weeks)
                is_season = weeks is None

        do_generate(
            sport=sport,
            weeks=weeks,
            is_season=is_season,
            season_id=selected_season_id,
            late_pick_threshold=args.late_pick_threshold,
            min_win_pct=args.min_win_pct,
            min_games_for_win_pct=args.min_games_win_pct,
            min_games=args.min_games,
            min_wins=args.min_wins,
        )
        return

    # Interactive mode
    sport = select_sport(title="Analysis Generator (Flat Table)", include_all=True)
    if not sport or sport == "exit":
        return

    if sport == "all":
        do_generate(sport=sport, weeks=None, is_season=True)
        return

    season_id = select_season(sport)
    if not season_id or season_id == "exit":
        return

    weeks, is_season = select_time_period(sport, season_id)
    if weeks is None and not is_season:
        return

    do_generate(sport=sport, weeks=weeks, is_season=is_season, season_id=season_id)


if __name__ == "__main__":
    main()
